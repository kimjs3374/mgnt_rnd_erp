# -*- coding: utf-8 -*-
"""① 증빙 판독 — 로컬 OCR 경로. **LLM 호출 0회.**

`extract.read_evidence()` 는 파일을 claude vision 에 던진다(실측 17초·$0.076).
이 모듈이 그 앞에 서서 **코드로 먼저 판독**하고, 산술로 검증된 결과가 나오면
LLM 을 아예 안 부른다. 못 뽑으면 조용히 물러나 기존 LLM 경로가 받는다.

반환 스키마는 `extract.read_evidence()` 와 **똑같다.** 뒤따르는
`resolve_direction()`·`verify_amounts()`·`classify()` 가 그대로 동작한다.

2026-09-03~04 실데이터 772파일 실측 근거 (01. 사전준비/프로토타입/ocr_selfmodel/):
  · 서식판별   2단 마커 + 카드전표·영수증 병합            85.7~91.1%
  · 회전보정   원본 쓰레기율>0.10 일 때만 OSD            +2.5p (무조건 걸면 역효과)
  · 금액추출   라벨 없이 산술 — s+v=t 이고 v≈s/10        커버리지 100%, 교차일치 94.1%
  · 사업자번호 가중모듈러 검증                            1자리 오독 100% 검출
  · OCR        tesseract psm4 (psm6 은 금액 폴백)        psm6 대비 서식 +37.6p

⚠ 거래 방향을 추측하지 않는다. 영수증·카드전표는 **가맹점이 공급자**라는 게 서식으로
   보장되므로 그때만 채우고, 그 외에는 비워서 `resolve_direction()` 이 「보류」하게 둔다.
   모델이 방향을 confidence 0.97 로 뒤집은 적이 있다(CLAUDE.md §5.2).
"""
from __future__ import annotations

import os
import re
import shutil
import subprocess
import tempfile
import zlib
from collections import Counter, defaultdict
from itertools import combinations

try:
    import numpy as np
    import cv2
    _HAS_CV = True
except ImportError:
    _HAS_CV = False

TESS = shutil.which("tesseract")
_HAS_TESS = bool(TESS)
OUR_BRN = re.sub(r"\D", "", os.environ.get("OUR_BRN", ""))

# ─────────────────────────────────────────────────────────────── 서식 판별
DOCTYPES = [
    ("견적서", [r"견\s*적\s*서", r"견\s*적\s*내\s*역", r"견\s*적\s*금\s*액"], [r"견\s*적"]),
    ("발주서", [r"발\s*주\s*서", r"주\s*문\s*서"], []),
    ("거래명세서", [r"거\s*래\s*명\s*세"], []),
    ("세금계산서", [r"세\s*금\s*계\s*산\s*서"], []),
    # ⚠ 「계산서」는 **면세** 거래다. 세금계산서와 다른 서식이고 부가세가 아예 없다.
    #   실측: 「전자계산서」(교육/도서 등 면세)를 세금계산서 마커가 못 잡아 weak 마커
    #   '승인번호' 가 걸리며 결제영수증으로 오분류됐다. 세금계산서보다 **뒤에** 둔다
    #   (앞에 두면 '세금계산서' 안의 '계산서' 가 먼저 걸린다).
    ("계산서", [r"전\s*자\s*계\s*산\s*서", r"(?<!세\s)(?<!세)계\s*산\s*서"], []),
    ("검수조서", [r"검\s*수\s*조\s*서", r"검\s*수\s*확\s*인"], [r"검\s*수"]),
    ("지출결의서", [r"지\s*출\s*결\s*의\s*서"], [r"지\s*출\s*결\s*의"]),
    # ★ 카드전표+영수증 병합. 필드라벨 4개가 결정적이었다(카드전표 recall 77.8%→88.9%)
    #
    # ⚠ 감열지 카드전표는 OCR 오독이 심하다. 실제로 걸린 것:
    #     '거라일시'(거래일시) · '|드총류'(카드종류) · '부가세물품가맥'(부가세물품가액)
    #   정확 일치만 보면 서식이 통째로 '미분류'가 된다 → 한 글자 흔들려도 잡히게 쓴다.
    ("결제영수증", [r"매\s*출\s*전\s*표", r"카\s*드\s*전\s*표", r"신용카드\s*매출", r"영\s*수\s*증",
                   r"거\s*[래라레러]\s*일\s*시",            # 거래일시
                   r"봉\s*사\s*료",
                   r"[카|]\s*[드匚]\s*[종총]\s*류",          # 카드종류
                   r"거\s*[래라레러]\s*유\s*형",             # 거래유형
                   r"부\s*가\s*세\s*물\s*품",                # 부가세물품가액
                   r"부\s*가\s*가\s*치\s*세?",               # 부가가치세
                   r"신\s*용\s*승\s*인", r"고\s*객\s*용"],
                  [r"가\s*맹\s*점", r"승\s*인\s*[번벤]\s*호", r"결\s*제\s*금\s*액",
                   r"할\s*부", r"일\s*시\s*불"]),
]
_PRIORITY = {t: i for i, (t, _, _) in enumerate(DOCTYPES)}
# extract.py 의 doc_type 어휘로 맞춘다
_DOC_MAP = {"세금계산서": "tax_invoice", "계산서": "tax_invoice",
            "거래명세서": "statement", "결제영수증": "receipt",
            "견적서": "quote", "발주서": "quote", "검수조서": "unknown",
            "지출결의서": "unknown"}


def classify_doc(text: str):
    if not text or len(re.sub(r"\s", "", text)) < 20:
        return None, 0.0, ""
    head = text[:4000]
    for tier, conf in ((1, 0.95), (2, 0.60)):
        best = None
        for entry in DOCTYPES:
            t = entry[0]
            for p in entry[tier]:
                m = re.search(p, head, re.I | re.M)
                if m:
                    cand = (m.start(), _PRIORITY[t], t, m.group(0))
                    if best is None or cand[:2] < best[:2]:
                        best = cand
                    break
        if best:
            return best[2], conf, best[3]
    return None, 0.0, ""


# ─────────────────────────────────────────────────────────────── 사업자번호
_W = (1, 3, 7, 1, 3, 7, 1, 3, 5)
# ⚠ OCR 은 하이픈을 온갖 것으로 읽는다. 실측: '119+86-81577' — `+` 때문에 못 잡아
#   거래 상대방 사업자번호를 통째로 놓쳤다. 구분자 후보를 넓히되, **체크섬이 최종
#   판정자**이므로 넓혀도 오탐이 늘지 않는다(무작위 통과율 9.99%).
_BRN_SEP = r"[-‒–—+~=*.,·:_|/\s]?"
BRN_RE = re.compile(rf"(?<!\d)(\d{{3}}){_BRN_SEP}\s*(\d{{2}}){_BRN_SEP}\s*(\d{{5}})(?!\d)")
# 칸에 한 자씩 흩어진 등록번호용. 하이픈 두 개를 **강제**해 승인번호·금액과 구분한다.
_D1 = r"\d[ \t]*"
_HY = r"[-‒–—~][ \t]*"
BRN_SPACED_RE = re.compile(
    rf"(?<!\d)(?:{_D1}){{3}}{_HY}(?:{_D1}){{2}}{_HY}(?:{_D1}){{4}}\d(?!\d)")


# 승인번호가 OCR 로 쪼개진 자리 — 10자리 뒤에 또 긴 숫자가 붙어 있으면 사업자번호가 아니다.
_APPROVAL_SPLIT = re.compile(r"(?<!\d)\d{10}[ \t]*\d{8,}")


def brn_valid(s: str) -> bool:
    s = re.sub(r"\D", "", s or "")
    if len(s) != 10:
        return False
    # 「0000000000」이 체크섬을 통과한다(가중합이 0이면 검증값도 0). 사업자번호일 리 없다.
    if len(set(s)) == 1:
        return False
    d = [int(c) for c in s]
    S = sum(w * x for w, x in zip(_W, d[:9])) + (d[8] * 5) // 10
    return (10 - S % 10) % 10 == d[9]


# 구분자 없는 10자리를 인정할 조건 — 같은 줄에 등록번호 라벨이 있어야 한다.
_BRN_LABEL = re.compile(r"등\s*록|사업자|번\s*호|사\s*업\s*자")
# 발급·결제 대행사가 자기 번호를 같이 찍는다. 거래 상대가 아니므로 후보에서 뺀다.
_AGENT_LINE = re.compile(r"대\s*행|발\s*급\s*업\s*무|UcessDI|Ucess|더존|나이스페이|"
                         r"이니시스|토스페이|KCP|PG\s*사|중개")


def find_brns(text: str) -> list[str]:
    out, seen = [], set()

    def _add(b: str) -> None:
        if b not in seen and brn_valid(b):
            seen.add(b)
            out.append(b)

    # ⚠ 줄 단위로 본다. BRN_RE 는 구분자를 생략할 수 있어(`?`) **하이픈 없는 긴 숫자열
    #   어디서든** 10자리를 집는다 — 승인번호 20240924410000080000 에서 2024092441 이
    #   나왔고, 상대 번호가 둘이 되어 매핑이 포기하면서 거래처가 비었다.
    #   구분자가 있으면 그대로 믿고, 없으면 등록번호 라벨이 같은 줄에 있을 때만 인정한다.
    for line in (text or "").split("\n"):
        # ★ 발급대행·결제대행사의 번호는 실재하지만 **거래 상대가 아니다.**
        #   실측: 「UcessDI ( 116-81-19477 )」가 거래처로 잡혔고,
        #        「발급업무 대행사업자 : (주)더존비즈온[134-81-08473]」도 마찬가지였다.
        if _AGENT_LINE.search(line):
            continue
        # ⚠ 라벨을 요구하면 **카드전표를 통째로 놓친다** — 「4088135513 (43)0629728292」
        #   처럼 라벨 없이 찍는다. 라벨 조건을 넣었던 이유(승인번호에서 번호를 지어냄)의
        #   범인은 슬라이딩 창이었고 이미 제거했다. 앞뒤 숫자 경계가 긴 숫자열을 막는다.
        for m in BRN_RE.finditer(line):
            # 승인번호 「2024092441 0000080000kp8w」에서 앞 10자리가 잘려 나오는 것을 막는다.
            if _APPROVAL_SPLIT.match(line, m.start()):
                continue
            _add("".join(m.groups()))
        # 칸에 한 자씩 흩어진 등록번호도 **같은 줄 판정 안에서** 본다.
        #   ⚠ 예전엔 이 루프가 문서 전체를 줄 구분 없이 훑어서, 위 대행사 제외가
        #     아무 소용이 없었다. 규칙이 두 군데 흩어지면 한쪽만 고치고 고쳤다고 믿는다.
        for m in BRN_SPACED_RE.finditer(line):
            _add(re.sub(r"\D", "", m.group()))

    # ★ 전자세금계산서는 등록번호를 **한 칸에 한 자씩** 찍는다. PDF 텍스트로 뽑으면
    #   칸 경계가 공백이 되어 「1 3 8 - 8 1 - 9 7 2 9 8」처럼 나온다. `\d{3}` 으로는
    #   못 잡는다 — 실측에서 상대 번호도 자사 번호도 통째로 놓쳤고, 거래처 사전에 있는
    #   업체(1388197298 (주)코리아사이언스)인데 조회조차 못 했다.
    #
    #   ⚠ 공백만 걷어내면 안 된다. 「1 3 8 - 8 1 - 9 7 2 9 8 408-81-68519」이 붙어
    #     「…97298408-81-…」이 되어 `(?!\d)` 경계에 걸려 **둘 다** 놓친다.
    #   그래서 숫자만 남긴 뒤 10자리 창을 앞에서부터 훑고, **체크섬이 통과하면 그만큼
    #   건너뛴다**(겹치는 오탐을 막는다). 무작위 통과율 9.99% 라 겹침만 막으면 안전하다.
    #
    #   ⚠ 하이픈류가 있는 줄에서만 한다. 전자세금계산서의 금액 칸
    #     「2024 04 29 5 1 1 4 0 0 0 …」처럼 구분자 없는 숫자 나열까지 훑으면
    #     10% 확률로 없는 번호를 지어낸다.
    # ★ 전자세금계산서는 등록번호를 칸에 한 자씩 찍는다 — 「1 3 8 - 8 1 - 9 7 2 9 8」.
    #   그래서 자릿수 사이 공백을 허용하되 **하이픈 두 개를 강제**한다.
    #
    #   ⚠ 예전에 「숫자만 남기고 10자리 창을 밀며 체크섬으로 거른다」를 썼다가 없는
    #     번호를 두 번 지어냈다: 승인번호 20240514-10240516-50184604 → 1410240516,
    #     일자+금액 2024/05/14 800,000 → 0514800000. 체크섬 통과율이 9.99% 라
    #     창을 여러 번 밀면 언젠가 걸린다. **조건을 좁히는 걸로는 못 막는다.**
    #     3-2-5 구조와 하이픈을 요구하면 승인번호도 금액도 자연히 빠진다.
    return out


def brn_paired_with_ours(text: str) -> str | None:
    """**우리 번호와 같은 줄**에 있는 상대 사업자번호. 세금계산서의 자리 규칙을 쓴다.

    세금계산서·계산서는 공급자와 공급받는자의 등록번호를 나란히 찍는다. 그 줄에 같이
    있는 번호가 거래처다 — 문서 하단의 발급대행사 번호와 확실히 구분된다.
    실측 배치: 「등록번호 220-87-95039  등록번호 408-81-68519」
    """
    if not OUR_BRN:
        return None
    for line in (text or "").split("\n"):
        if _AGENT_LINE.search(line):
            continue
        bs = []
        for m in BRN_RE.finditer(line):
            if re.search(r"[-‒–—~]", m.group()) or _BRN_LABEL.search(line):
                b = "".join(m.groups())
                if brn_valid(b) and b not in bs:
                    bs.append(b)
        for m in BRN_SPACED_RE.finditer(line):
            b = re.sub(r"\D", "", m.group())
            if brn_valid(b) and b not in bs:
                bs.append(b)
        if OUR_BRN in bs:
            others = [b for b in bs if b != OUR_BRN]
            if len(others) == 1:
                return others[0]
    return None


# ─────────────────────────────────────────────────────────────── 일자
_DATE_PATS = [
    (re.compile(r"(20\d{2})\s*[-./년]\s*(\d{1,2})\s*[-./월]\s*(\d{1,2})"), 2000),
    (re.compile(r"(?<!\d)(\d{2})\s*[-./]\s*(\d{1,2})\s*[-./]\s*(\d{1,2})(?!\d)"), 2000),
    # ★ 전자세금계산서 작성일자 칸: 구분자 없이 「2024 04 29」로 흩어진다.
    #   구분자가 없어 느슨하므로 **맨 뒤에 둔다** — 위 두 패턴이 먼저 잡으면 그걸 쓴다.
    #   년/월/일 범위 검증은 find_date 가 그대로 한다.
    (re.compile(r"(?<!\d)(20\d{2})\s+(\d{1,2})\s+(\d{1,2})(?!\d)"), 2000),
]


# 전자세금계산서 승인번호는 **앞 8자리가 작성일자**다 — 20240924410000080000.
# 문서에서 가장 안정적인 날짜 신호라 먼저 본다.
_APPROVAL_RE = re.compile(r"(?<!\d)(20\d{6})[-\s]?[\dA-Za-z]{8,}")
# 「국세청 고시 제2013-17호, 2013.4.1」 같은 **법령 문구의 날짜**를 집으면 안 된다.
_LAW_LINE = re.compile(r"국세청|고시|법적|효력|미\s*전송|약관|조\s*항")


def find_date(text: str) -> str | None:
    # ① 승인번호에 박힌 작성일자
    for m in _APPROVAL_RE.finditer(text or ""):
        s = m.group(1)
        y, mo, d = int(s[:4]), int(s[4:6]), int(s[6:8])
        if 2000 <= y <= 2100 and 1 <= mo <= 12 and 1 <= d <= 31:
            return f"{y:04d}-{mo:02d}-{d:02d}"
    # ② 그 밖의 날짜 — 법령 문구 줄은 뺀다
    text = "\n".join(ln for ln in (text or "").split("\n") if not _LAW_LINE.search(ln))
    for pat, base in _DATE_PATS:
        for m in pat.finditer(text or ""):
            y, mo, d = m.groups()
            y = int(y) + (base if len(y) == 2 else 0)
            mo, d = int(mo), int(d)
            if 2000 <= y <= 2100 and 1 <= mo <= 12 and 1 <= d <= 31:
                return f"{y:04d}-{mo:02d}-{d:02d}"
    return None


# ─────────────────────────────────────────────────────────────── 금액 (라벨 없이)
# ⚠ OCR 은 천단위 구분자를 `,` 로만 주지 않는다. 실제로 걸린 것:
#     '70.' '000'   <- "70,000" 이 마침표 + 공백으로 쪼개져 두 토큰이 됐다
#     '54.900'      <- 쉼표를 마침표로 읽었다
#   그래서 (1) `.` 도 구분자로 받고 (2) **같은 행에서 쪼개진 숫자 토큰을 다시 붙인다.**
#   이걸 안 하면 카드전표에서 금액 후보가 0개가 되어 산술검증이 통째로 실패한다.
_SEP = r"[,.·]"
_AMT_HEAD = re.compile(rf"^[₩\\]?\(?-?\d{{1,3}}{_SEP}?$")          # '70.'  '1,'  '142'
_AMT_TAIL = re.compile(rf"^\d{{3}}{_SEP}?\)?원?$")                  # '000'  '646,'
_AMT_WHOLE = re.compile(rf"^[₩\\]?\(?-?\d{{1,3}}(?:{_SEP}\d{{3}})+\)?원?$")  # '1,394,646'
_AMT_BARE = re.compile(r"^[₩\\]?\(?-?\d{3,9}\)?원?$")               # '70000'
_NOT_AMT = re.compile(r"번호|No|NO|승인|카드|사업자|전화|TEL|Tel|일시|날짜|일자|시간|"
                      r"POS|가맹점|단말|거래번호|주소|우편")
_DATE_LIKE = re.compile(r"^(19|20)\d{2}$|^\d{6}$|^\d{8}$")
# 숫자 오른쪽에 이런 단위가 붙으면 금액이 아니라 **주소**다.
_ADDR_UNIT = re.compile(r"^(호|동|층|번지|가|호실|호,|동,)$")


def _digits(t: str) -> str:
    return re.sub(r"\D", "", t or "")


def _amount(tok: str, *, trusted: bool = False):
    """단일 토큰 금액. trusted=True 면 구분자가 확인된 것이라 날짜 필터를 건너뛴다."""
    t = (tok or "").strip()
    if _AMT_WHOLE.match(t):
        trusted = True
    elif not _AMT_BARE.match(t):
        return None
    s = _digits(t)
    if not s or (not trusted and _DATE_LIKE.match(s)):
        return None
    v = int(s)
    return v if 100 <= v <= 100_000_000 else None


def _merge_row(ws):
    """같은 행에서 OCR 이 쪼갠 숫자 토큰을 다시 붙인다. ('70.','000') -> 70000"""
    out, i = [], 0
    ws = sorted(ws, key=lambda w: w[0])
    while i < len(ws):
        x0, x1, top, bot, cf, txt = ws[i]
        t = txt.strip()
        if _AMT_HEAD.match(t):
            parts, j, last_x1 = [t], i + 1, x1
            h = max(6.0, bot - top)
            while j < len(ws):
                nx0, nx1, ntop, nbot, ncf, ntxt = ws[j]
                # 붙어 있어야 한 숫자다. 멀면 다른 칸이다.
                if nx0 - last_x1 > h * 1.2 or not _AMT_TAIL.match(ntxt.strip()):
                    break
                parts.append(ntxt.strip())
                last_x1 = nx1
                j += 1
            if len(parts) > 1:
                s = _digits("".join(parts))
                if 3 <= len(s) <= 12:
                    out.append(dict(v=int(s), x0=x0, y=(top + bot) / 2,
                                    h=bot - top, sep=True, idx=i))
                    i = j
                    continue
        v = _amount(t)
        if v is not None:
            out.append(dict(v=v, x0=x0, y=(top + bot) / 2, h=bot - top,
                            sep=bool(re.search(_SEP, t)), idx=i))
        i += 1
    return out


def group_rows(words, tol_ratio: float = 0.6):
    """y 좌표로 행을 묶는다.

    ⚠ **토큰 자신의 높이로 나눠 행 키를 만들면 안 된다.** 실제로 걸린 것:
       같은 줄의 '70.'(h=54) 과 '000'(h=32) 이 각각 키 22 와 37 로 갈라져
       숫자 병합이 통째로 실패했다. 문서 전체의 **중앙값 높이**를 기준으로 묶는다.
    """
    if not words:
        return []
    hs = sorted(max(1.0, w[3] - w[2]) for w in words)
    med = hs[len(hs) // 2]
    tol = max(4.0, med * tol_ratio)
    ws = sorted(words, key=lambda w: (w[2] + w[3]) / 2)
    rows, cur, cur_y = [], [], None
    for w in ws:
        y = (w[2] + w[3]) / 2
        if cur_y is None or abs(y - cur_y) <= tol:
            cur.append(w)
            cur_y = y if cur_y is None else (cur_y * (len(cur) - 1) + y) / len(cur)
        else:
            rows.append(cur)
            cur, cur_y = [w], y
    if cur:
        rows.append(cur)
    return rows


def _amt_tokens(words):
    out = []
    for ws in group_rows(words):
        ws_sorted = sorted(ws, key=lambda w: w[0])
        for cand in _merge_row(ws):
            # 같은 행 왼쪽 3토큰에 '번호/승인/카드' 류가 있으면 금액이 아니다
            k = cand.pop("idx")
            if any(_NOT_AMT.search(ws_sorted[j][5])
                   for j in range(max(0, k - 3), k)):
                continue
            # ★ 오른쪽에 주소 단위가 붙으면 금액이 아니다 — 「102 동」 「1106 호」.
            #   실측: 케미칼뱅크 세금계산서에서 주소의 호수 1106 이 합계로 잡혔고,
            #   같은 주소의 102 와 11 배 관계를 이뤄 「검산 통과」까지 했다.
            if any(_ADDR_UNIT.match((ws_sorted[j][5] or "").strip())
                   for j in range(k + 1, min(len(ws_sorted), k + 3))):
                continue
            cand["comma"] = cand.pop("sep")
            out.append(cand)
    return out


# ── 라벨 앵커: 문서가 직접 「합계금액」이라고 말하는 자리 ────────────────────
# ⚠ 「합계」 단독은 넣지 않는다 — 「합계 968,000 96,800」처럼 공급가액·세액 소계 행이
#   같은 말을 쓴다. 최종 금액을 뜻하는 말만 앵커로 삼는다.
# ⚠ OCR 은 「합계금액」을 「함 계 금액」으로 읽는다(실측, 포스코 세금계산서).
#   글자 오독을 관용하되 「합계」 단독은 여전히 넣지 않는다 — 소계 행과 섞인다.
_TOTAL_LAB = re.compile(r"[합함항][계게]금[액엑]|총결제금[액엑]|총[합함][계게]|청구금[액엑]|"
                        r"승인금[액엑]|결제금[액엑]|받을금[액엑]")


def _total_by_label(words) -> int | None:
    """「합계금액」 라벨을 찾아 그 값을 집는다. 산술 추론보다 **문서가 직접 말한 값**이 낫다.

    측정으로 확인한 배치는 둘이다:
      · 같은 행 오른쪽   「합계금액 ￦ 1,064,800 (부가세포함)」
      · 다음 행         「합 계 금 액 | 현금 | 수표 …」 아래에 「1,443,750 …」
    """
    if not words:
        return None
    rows = group_rows(words)
    for ri, ws in enumerate(rows):
        ws = sorted(ws, key=lambda w: w[0])
        if not _TOTAL_LAB.search(re.sub(r"\s", "", "".join(w[5] for w in ws))):
            continue
        # ① 같은 행의 금액 중 가장 큰 값. 현금·수표·어음 0 은 _amount 가 걸러낸다.
        #    ⚠ **_merge_row 를 쓴다.** OCR 은 금액을 「207,」+「200」처럼 쪼개는데,
        #      원본 토큰만 보면 그걸 못 읽어 라벨이 있는데도 앵커가 실패한다(실측).
        vals = [c["v"] for c in _merge_row(ws)]
        if vals:
            return max(vals)
        # ② 라벨만 있는 헤더 행이면 다음 행에서 **라벨 열 근처**의 값을 본다.
        if ri + 1 < len(rows):
            lab_x = ws[0][0]
            span = max(240, (ws[-1][0] - ws[0][0]) // 2)
            near = [c["v"] for c in _merge_row(rows[ri + 1])
                    if abs(c["x0"] - lab_x) <= span]
            if near:
                return max(near)
    return None


def extract_amounts(words):
    """s+v=t 이고 v≈s/10 이면 라벨이 없어도 세 값이 동시에 확정된다."""
    toks = _amt_tokens(words)
    vals = [t["v"] for t in toks]
    if not vals:
        return dict(합계=None, 공급가액=None, 세액=None, 방법="없음", 산술검증=False)
    vs = set(vals)

    # ★ S0 라벨 앵커 — 문서가 「합계금액」이라고 직접 말한 값이 산술 추론보다 낫다.
    #   실측: 합계 오답 4건이 전부 라벨이 있는데도 부가세·공급가액·첫 품목을 집은 경우였다.
    #   라벨값이 공급가액+세액과도 맞으면 검산까지 된 것으로 본다.
    t0 = _total_by_label(words)
    if t0:
        for s in sorted(vs, reverse=True):
            v = t0 - s
            if v in vs and v > 0 and abs(v * 10 - s) <= max(2, s * 0.02):
                return dict(합계=t0, 공급가액=s, 세액=v,
                            방법="S0-합계금액라벨(검산됨)", 산술검증=True)
        v0 = round(t0 / 11)
        공급, 세 = (t0 - v0, v0) if (v0 in vs or (t0 - v0) in vs) else (None, None)
        return dict(합계=t0, 공급가액=공급, 세액=세,
                    방법="S0-합계금액라벨", 산술검증=False)
    for s in sorted(vs):                                   # S1 삼중항
        for dv in (0, -1, 1):
            v = round(s / 10) + dv
            if v >= 10 and v in vs and (s + v) in vs:
                return dict(합계=s + v, 공급가액=s, 세액=v,
                            방법="S1-산술삼중항", 산술검증=True)
    # S1b 공급가액·부가세 쌍 — 합계가 안 읽혀도 10% 관계가 확인되면 더해서 만든다.
    #   실측: 감열지 카드전표에서 '70.000'·'7.000' 은 읽혔는데 합계 '77.000' 만 뭉갰다.
    #   두 값의 10:1 관계 자체가 근거이므로 합계를 계산해도 추측이 아니다.
    best = None
    for s in sorted(vs, reverse=True):
        for dv in (0, -1, 1):
            v = round(s / 10) + dv
            if v >= 100 and v in vs and s != v:
                if best is None or s > best[0]:
                    best = (s, v)
                break
    if best:
        s, v = best
        return dict(합계=s + v, 공급가액=s, 세액=v,
                    방법="S1b-공급가액·부가세쌍(합계는 계산)", 산술검증=True)
    for t in sorted(vs, reverse=True):                     # S2 11분할
        v0 = round(t / 11)
        for v in (v0, v0 + 1, v0 - 1):
            if v >= 10 and v in vs:
                # ⚠ **산술검증=False.** t 와 v 가 문서에 있기만 하면 통과하므로 근거가 약하다.
                #   실측: 주소의 「102 동 1106 호」가 우연히 11 배 관계를 이뤄 합계 1,106원이
                #   「검산 통과」로 나왔다. 자기가 고른 값으로 자기를 검산한 셈이다.
                #   검산이라고 말할 수 있는 건 S1/S1b(s+v=t 이고 v≈s/10) 뿐이다.
                return dict(합계=t, 공급가액=t - v, 세액=v,
                            방법="S2-11분할(미검산)", 산술검증=False)
    # S3 품목합산 — 행 묶기는 중앙값 높이 기준(토큰 자기 높이로 나누면 같은 줄이 갈라진다)
    hs = sorted(max(1.0, t["h"]) for t in toks)
    med = hs[len(hs) // 2]
    tol = max(4.0, med * 0.6)
    rows, cur, cur_y = [], [], None
    for t in sorted(toks, key=lambda t: t["y"]):
        if cur_y is None or abs(t["y"] - cur_y) <= tol:
            cur.append(t)
            cur_y = t["y"] if cur_y is None else (cur_y * (len(cur) - 1) + t["y"]) / len(cur)
        else:
            rows.append(cur)
            cur, cur_y = [t], t["y"]
    if cur:
        rows.append(cur)
    lines = []
    for ts in rows:
        vv = [t["v"] for t in sorted(ts, key=lambda t: t["x0"])]
        for a, b, c in combinations(vv, 3):
            if a * b == c:
                lines.append(c)
                break
    if lines and sum(lines) in vs:
        return dict(합계=sum(lines), 공급가액=None, 세액=None,
                    방법=f"S3-품목합산({len(lines)}행)", 산술검증=True)
    # S5 면세 반복일치 — 부가세가 **없는** 문서(계산서·면세영수증)용.
    #   10% 관계가 원리적으로 성립하지 않으므로 다른 근거를 쓴다: 계산서는 합계를
    #   공급가액·품목금액·영수금액 칸에 **여러 번** 인쇄한다. 3회 이상 같은 값이
    #   반복되면 서식 구조가 만든 일치이지 우연이 아니다.
    #   실측: 전자계산서(면세) 648,000 이 4회 반복 — S1~S4 가 전부 못 잡던 문서다.
    cnt = Counter(vals)
    rep = [v for v, c in cnt.items() if c >= 3]
    if rep:
        t = max(rep)
        return dict(합계=t, 공급가액=t, 세액=0,
                    방법=f"S5-면세반복일치({cnt[t]}회)", 산술검증=True)

    comma = [t["v"] for t in toks if t["comma"]]           # S4 최대금액 (근거 없음)
    return dict(합계=max(comma) if comma else max(vals), 공급가액=None, 세액=None,
                방법="S4-최대금액", 산술검증=False)


_KOR = re.compile(r"[가-힣]{2,}")


# 품목명이 아닌 것들 — 서식 라벨·머리글이 품목으로 새어 들어온다
_NOT_ITEM = re.compile(
    r"^(합계|소계|총액|총계|공급가액|부가세|세액|과세|면세|금액|단가|수량|품목|품명|상품명|"
    r"거래처|상호|주소|전화|대표|사업자|일자|날짜|승인|카드|가맹점|매출|영수증|전표|"
    r"결제|현금|신용|할부|일시불|받을금액|청구금액|봉사료|담당|비고|고객|매장|점포|"
    r"물품|가액|부가가치|거래|번호|잔액|포인트|적립|할인)")

# 주소는 품목이 아니다. 실측: 카드전표에서 '광주 광주 남구 남구' 가 품목으로 잡혀
# 자체 모델 확신도를 0.608 -> 0.552 로 **떨어뜨렸다**. 틀린 품목은 안 주느니만 못하다.
_ADDR = re.compile(
    r"^(서울|부산|대구|인천|광주|대전|울산|세종|경기|강원|충북|충남|전북|전남|경북|경남|제주)"
    r"|(시|군|구|읍|면|동|리|로|길|번지|가)$")


# 세금계산서·계산서·거래명세서의 품목 표는 구조가 고정돼 있다.
#   헤더:  월 일 품목명 규격 수량 단가 공급가액 세액 비고
#   행:    09 25 배터리신뢰성시험            22,670,000  2,267,000
#   종료:  합계금액 현금 수표 어음 외상미수금
# **품목이 비목을 정한다.** 거래처가 정하는 게 아니다 — 같은 거래처에서 장비도 사고
# 재료도 산다. 그래서 이 표를 제대로 읽는 것이 비목 분류의 출발점이다.
_ITEM_HDR = re.compile(r"(품\s*목\s*명?|품\s*명|상\s*품\s*명)")
_ITEM_HDR2 = re.compile(r"(수\s*량|단\s*가|규\s*격|공\s*급\s*가\s*액)")
_ITEM_END = re.compile(r"합\s*계\s*금\s*액|이\s*금\s*액|외\s*상\s*미\s*수\s*금|"
                       r"소\s*계|공\s*급\s*가\s*액\s*계|비\s*고\s*$")


def extract_items_from_text(text: str) -> list[str]:
    """품목 표에서 품목명만 뽑는다. 금액·수량·날짜는 걷어낸다."""
    lines = [l for l in (text or "").split("\n")]
    start = None
    for i, ln in enumerate(lines):
        if _ITEM_HDR.search(ln) and _ITEM_HDR2.search(ln):
            start = i + 1
            break
    if start is None:
        return []
    out = []
    for ln in lines[start:start + 20]:
        if _ITEM_END.search(ln):
            break
        s = re.sub(r"^\s*\d{1,2}\s+\d{1,2}\s+", " ", ln)   # 앞의 월 일
        s = re.sub(r"[\d,]{3,}", " ", s)                    # 금액·수량
        s = re.sub(r"[^\w가-힣()%/.\-\s]", " ", s)
        s = re.sub(r"\s+", " ", s).strip(" .,-|/")
        flat = re.sub(r"\s", "", s)
        if len(flat) < 2 or _NOT_ITEM.match(flat) or _ADDR.search(flat):
            continue
        # 한글이나 영문이 있어야 품목명이다
        if not re.search(r"[가-힣A-Za-z]{2,}", s):
            continue
        out.append(s[:60])
    return out[:12]


def extract_items(words):
    """같은 행에 한글 품명 + 금액이 있으면 품목 후보로 본다.

    ②비목 분류(classify)가 품목명을 먹으므로 거칠어도 이름은 넣어준다.

    ⚠ **금액은 일부러 None 으로 둔다.** 행에서 집은 숫자가 단가인지 금액인지 소계인지
       코드가 구분 못 한다. 실측에서 거래명세표의 항목합이 6,508,800 으로 잡혀
       합계 2,237,400 과 어긋나 `verify_amounts()` 가 가짜 「항목합_불일치」를 냈다.
       금액을 비우면 그 검산이 이 항목들을 건너뛴다 — **틀린 숫자를 주느니 안 준다.**
    """
    items, seen = [], set()
    for ws in group_rows(words):
        ws = sorted(ws, key=lambda w: w[0])
        names = [w[5].strip() for w in ws
                 if _KOR.fullmatch(w[5].strip())
                 and not _NOT_AMT.search(w[5]) and not _NOT_ITEM.match(w[5].strip())
                 and not _ADDR.search(w[5].strip())]
        if not names or not any(_amount(w[5]) for w in ws):
            continue
        nm = " ".join(names)[:40]
        if nm in seen or not _looks_like_item(nm):
            continue
        seen.add(nm)
        items.append({"품목명": nm, "수량": None, "금액": None, "confidence": 0.45,
                      "note": "코드 추출(품명만 · 금액은 합계로 검산)"})
    return items[:12]


# 영수증 서식의 라벨·안내문이 품목으로 새어나온다. 실측으로 걸린 것:
#   「타계 금액」「가세 과 세 물 품 가오」「신 용 승 인 정 보」「부가세」
# 이런 조각이 비목 모델에 들어가 「칼국수집 → 재료비」 오답을 만들었다.
_ITEM_NOISE = re.compile(
    r"금액|가액|과세|면세|부가|세액|합계|소계|총액|승인|정보|카드|신용|현금|결제|"
    r"영수|증빙|매출|전표|가맹|점명|사업자|번호|일시|일자|시간|주소|전화|대표|"
    r"할부|일시불|봉사료|포인트|적립|잔액|거스름|받을|청구|구매|판매|공급|"
    r"단말|매입|승인|취소|반품|교환|안내|문의|감사|고객|보관|용도")


def _looks_like_item(nm: str) -> bool:
    """품목명처럼 보이는가. 라벨·안내문이면 버린다 — 틀린 품목은 안 주느니만 못하다."""
    flat = re.sub(r"\s", "", nm)
    if len(flat) < 2 or len(flat) > 30:
        return False
    if _ITEM_NOISE.search(flat):
        return False
    # 한글 2자 이상이 이어져야 품명이다 ('ㄱ oa' 같은 조각 배제)
    return bool(re.search(r"[가-힣]{2,}", flat))


# ─────────────────────────────────────────────────────────────── 상호
_NAME_PATS = [
    re.compile(r"가\s*맹\s*점\s*(?:명)?\s*[:：]\s*([^\n:：]{2,30})"),
    re.compile(r"상\s*호\s*(?:명)?\s*[:：]\s*([^\n:：]{2,30})"),
    re.compile(r"매\s*장\s*명\s*[:：]?\s*([^\n:：]{2,30})"),
]
# ⚠ 표 서식 카드전표는 라벨과 값이 **다른 줄**에 있고 콜론이 없다. 실측:
#     가 맹 점 명
#     ‘BSR 깐 깐 한 족 발 ( 떼 이 >)
#   콜론을 요구하는 패턴만 쓰면 가맹점을 통째로 놓친다 — 카드전표는 가맹점이
#   비목의 **유일한** 근거이므로(품목이 인쇄되지 않는다) 이걸 놓치면 분류가 불가능하다.
_NAME_LABEL_LINE = re.compile(r"^\s*(가\s*맹\s*점\s*명?|상\s*호|매\s*장\s*명)\s*[:：]?\s*$")


# 우리 상호. 좌우 블록이 뒤섞였을 때 **우리를 거래처로 잡는 사고**를 막는 마지막 방어선.
OUR_NAME = os.environ.get("OUR_NAME", "매그나텍")
_NAME_LABEL = re.compile(r"상\s*호|법\s*인\s*명|성\s*명|등\s*록|번\s*호|사\s*업\s*장|주\s*소|"
                         r"업\s*태|종\s*목|이\s*메\s*일|Email|종사업|공\s*급|받\s*는\s*자|대\s*표")
_NAME_OK = re.compile(r"[가-힣A-Za-z]{2,}")


def find_vendor_name(words, text: str, their_brn: str | None) -> str | None:
    """상대 사업자번호가 찍힌 **자리**를 기준으로 상호를 집는다.

    ⚠ 텍스트 순서로 집으면 안 된다. 좌우 블록이 한 줄로 펼쳐져 「김용민 공 ㈜매그나텍」처럼
      두 회사가 섞인다 — 우리 회사가 거래처로 들어갈 뻔했다(실측).
    """
    if not words or not their_brn:
        return None
    rows = group_rows(words)
    d = re.sub(r"\D", "", their_brn)
    for ri, ws in enumerate(rows):
        ws = sorted(ws, key=lambda w: w[0])
        joined = re.sub(r"[^\d]", "", "".join(w[5] for w in ws))
        if d not in joined:
            continue
        # 상대 번호 토큰의 x 위치. 숫자 조각이 흩어져 있으면 그 구간의 가운데.
        xs = [w[0] for w in ws if re.search(r"\d", w[5])]
        if not xs:
            continue
        # 우리 번호가 같은 줄에 있으면 좌우가 갈린다 — 우리 쪽 절반을 잘라낸다.
        our = re.sub(r"\D", "", OUR_BRN or "")
        mid = None
        if our and our in joined:
            acc, our_x, their_x = "", None, None
            for w in ws:
                if not re.search(r"\d", w[5]):
                    continue
                acc += re.sub(r"\D", "", w[5])
                if our_x is None and our in acc:
                    our_x = w[0]
                if their_x is None and d in acc:
                    their_x = w[0]
            if our_x is not None and their_x is not None and our_x != their_x:
                mid = (our_x + their_x) / 2
                왼쪽 = their_x < our_x
        # 아래 2줄에서 상호 후보를 모은다
        for nx in (ri + 1, ri + 2):
            if nx >= len(rows):
                break
            cand = []
            for w in sorted(rows[nx], key=lambda w: w[0]):
                if mid is not None and ((w[0] > mid) if 왼쪽 else (w[0] < mid)):
                    continue                    # 우리 블록 쪽은 보지 않는다
                t = (w[5] or "").strip()
                if not t or _NAME_LABEL.fullmatch(re.sub(r"\s", "", t)):
                    continue
                if re.fullmatch(r"[\d\-‒–—~.,()|/]+", t):
                    continue
                cand.append(t)
            nm = _clean_vendor_name(" ".join(cand))
            if nm:
                return nm
    return None


# 세로쓰기 라벨이 한 글자씩 떨어져 나온다 — 「공 급 자」 「받 는 자」 「상 호」 「성 명」.
_LABEL_TOK = {"상", "호", "성", "명", "법", "인", "등", "록", "번", "공", "급", "자",
              "받", "는", "사", "업", "장", "주", "소", "업", "태", "종", "목", "대", "표",
              "상호", "성명", "법인명", "등록", "번호", "공급자", "받는자", "사업장",
              "주소", "업태", "종목", "대표", "대표자", "이메일", "종사업", "장번호",
              # 「(주)」「(법인명)」이 칸 경계로 쪼개져 들어온다
              "법인", "인명", "귀하", "귀중", "보관용", "일자", "품목"}
# 「(주)」가 칸 경계로 쪼개져 「주)」 「(주」 처럼 붙어 들어온다. 이 조각만 떼어낸다.
#   ⚠ 「주」를 라벨 목록에 넣으면 안 된다 — 「주) 후성」이 「후성」이 되고 그게 길이
#     제한에 걸려 탈락하면서 **다음 줄의 주소**를 집었다(실측).
_CORP_FRAG = re.compile(r"^[(（]?[주유][)）]?$|^[(（]?[주유][)）]")


# 대표자명 판별 — **성씨를 반드시 요구한다.** 성씨를 선택으로 두면 아무 한글 2~3자나
# 잘려나가 상호의 일부까지 지운다(「(주)한빛 과학」의 「과학」).
_SURNAME = ("김이박최정강조윤장임한오서신권황안송전홍고문양손배백허유남심노하곽성차주우구"
            "나민진지엄채원천방공현함변염여추도소석선설마길연위표명기반왕금옥육인맹제모탁국")
_PERSON = re.compile(rf"[{_SURNAME}][가-힣]{{1,2}}")


def _clean_vendor_name(raw: str) -> str | None:
    """상호로 쓸 수 있는가. **의심스러우면 비운다.**

    실측에서 걸린 것들: 「성 상호 성」(라벨 조각) · 「김용민 공 ㈜매그나텍」(대표자명+우리 회사)
    · 「' 주식회사 ㄱㄱ OM Al Ad oo o4」(OCR 쓰레기). 이런 이름이 원장에 들어가면
    정산에서 반려되고, 우리 회사가 거래처로 들어가면 그건 사고다.
    비어 있으면 사업자번호로 사전을 조회하고, 없으면 사람이 한 번 적는다 — 길이 있다.
    """
    toks = [t for t in re.split(r"[\s·,.|/]+", raw or "") if t]
    toks = [re.sub(r"^[()\[\]]+|[()\[\]]+$", "", t) for t in toks]
    toks = [_CORP_FRAG.sub("", t) for t in toks]
    toks = [t for t in toks if t and t not in _LABEL_TOK]
    # ★ 상호 칸 옆이 성명 칸이라 대표자명이 딸려 온다 — 「선일상사 김광열」 「주)대송 박중하」.
    while len(toks) >= 2 and _PERSON.fullmatch(toks[-1]):
        toks.pop()
    nm = re.sub(r"\s+", " ", " ".join(toks)).strip(" ·,.-|/()")
    flat = re.sub(r"\s", "", nm)
    # 「후성」·「엔켐」처럼 **두 글자 상호**가 실제로 있다. 한글 2자 조건이 쓰레기를 막는다.
    if len(flat) < 2:
        return None
    # 우리 회사가 섞였으면 쓰지 않는다. 좌우 블록이 뒤엉킨 것이다.
    if OUR_NAME and OUR_NAME.replace(" ", "") in flat:
        return None
    # 한글 자모만 남은 조각(ㄱㄱ, ㅁㄴ)이 있으면 OCR 이 뭉갠 것이다
    if re.search(r"[ㄱ-ㅎㅏ-ㅣ]", flat):
        return None
    # 한글 상호에 라틴 문자가 섞이면 OCR 오독이다(영문 상호는 한글이 없다)
    han = len(re.findall(r"[가-힣]", flat))
    lat = len(re.findall(r"[A-Za-z]", flat))
    if han and lat and lat > han * 0.4:
        return None
    if han < 2:
        # 한글이 없다 = 영문 상호이거나 OCR 쓰레기. 진짜 영문 상호는 토큰이 **대문자로
        # 시작**한다(Adobe Systems Software Ireland Ltd). 쓰레기는 소문자 조각이 섞인다
        # (실측: 「jfmon Pecere AA Sl xo au」).
        words_ = [w for w in re.split(r"\s+", nm) if re.search(r"[A-Za-z]", w)]
        if len(lat and words_ or []) < 1 or lat < 3:
            return None
        if not all(w[0].isupper() for w in words_):
            return None
    return nm[:60]


def find_merchant(text: str) -> str | None:
    """가맹점·상호 한 개를 뽑는다.

    ⚠ 세금계산서·거래명세서처럼 **양쪽 상호가 다 찍힌 서식에서는 쓰지 않는다.**
       어느 쪽이 공급자인지 좌표 없이 알 수 없고, 잘못 붙이면 거래처가 뒤집힌다 —
       모델이 그걸 confidence 0.97 로 틀린 적이 있다(CLAUDE.md §5.2).
       그래서 호출부에서 결제영수증일 때만 부른다.
    """
    for pat in _NAME_PATS:
        m = pat.search(text or "")
        if m:
            nm = _clean_name(m.group(1))
            if nm:
                return nm
    # 라벨만 있는 줄 다음에서 찾는다 (표 서식 카드전표)
    lines = (text or "").split("\n")
    for i, ln in enumerate(lines):
        if not _NAME_LABEL_LINE.match(ln):
            continue
        for j in range(i + 1, min(i + 4, len(lines))):
            cand = re.sub(r"[^\w가-힣()\s]", " ", lines[j])
            cand = re.sub(r"\s+", " ", cand).strip()
            if not cand or _NAME_LABEL_LINE.match(lines[j]):
                continue
            nm = _clean_name(cand)
            if nm:
                return nm
    return None


# 라벨어 자체를 상호로 반환하면 안 된다. 실측: 세금계산서에서 「상 호」가 거래처명으로
# 잡혔고, 그 쓰레기 이름이 비목 모델 입력까지 오염시켰다.
_LABEL_WORDS = re.compile(
    r"^(상호|법인명|성명|이름|등록번호|사업자|사업장|주소|업태|종목|전화|이메일|"
    r"공급자|공급받는자|받는자|대표|대표자|비고|품목|수량|단가|금액|합계|작성일자)$")


def _clean_name(s: str) -> str | None:
    nm = re.sub(r"\s+", " ", s or "").strip(" .,-|:()")
    nm = re.sub(r"\s*(대표|사업자|전화|TEL|주소|성명|업태|종목|이메일).*$", "", nm).strip()
    flat = re.sub(r"\s", "", nm)
    if len(flat) < 2 or _LABEL_WORDS.match(flat):
        return None
    # OCR 쓰레기 걸러내기 — 한글 비중이 너무 낮으면 이름으로 안 본다
    kor = sum(1 for c in flat if "가" <= c <= "힣")
    if kor < 2 or kor / len(flat) < 0.5:
        return None
    return nm[:30]


_NAME_NEAR = re.compile(r"(?:상\s*호|법\s*인\s*명|성\s*명)\s*[:：]?\s*([^\n:：]{2,30})")


def find_name_in_brn_row(words, brn: str) -> str | None:
    """사업자번호와 **같은 줄**에 있는 한글 상호를 집는다.

    카드전표는 상호를 라벨 없이 왼쪽에, 사업자번호를 오른쪽에 같은 줄로 찍는다.
        자금성                    308-38-01037
    라벨(`상호:`)이 없어 패턴 검색이 실패하던 케이스다. 줄을 기준으로 잡으면 잡힌다.
    """
    d = _digits(brn)
    if not d or not words:
        return None
    for row in group_rows(words):
        row = sorted(row, key=lambda w: w[0])
        if not any(_digits(w[5]) and _digits(w[5]) in d for w in row):
            continue
        if d not in _digits("".join(w[5] for w in row)):
            continue
        names = [w[5].strip() for w in row
                 if _KOR.fullmatch(w[5].strip()) and not _NOT_ITEM.match(w[5].strip())]
        if names:
            return " ".join(names)[:30]
    return None


def find_name_near_brn(text: str, brn: str) -> str | None:
    """**이미 확정된 사업자번호 위치**를 기준으로 그 근처의 상호를 집는다.

    양쪽 상호가 다 찍힌 서식(세금계산서·거래명세서)에서 「어느 쪽이 공급자인가」를
    추측하지 않기 위한 방법이다. 방향은 사업자번호로 이미 코드가 정했고, 이름은
    그 번호 **주변에서만** 찾는다. 번호를 못 찾으면 아무것도 반환하지 않는다.
    """
    if not brn or not text:
        return None
    d = re.sub(r"\D", "", brn)
    # 문서에는 하이픈이 있을 수 있다 — 숫자만 남긴 사본에서 위치를 찾고 원문으로 되돌린다
    flat, idx_map = [], []
    for i, ch in enumerate(text):
        if ch.isdigit():
            flat.append(ch)
            idx_map.append(i)
    pos = "".join(flat).find(d)
    if pos < 0:
        return None
    at = idx_map[pos]
    window = text[max(0, at - 160): at + 160]
    best = None
    for m in _NAME_NEAR.finditer(window):
        nm = _clean_name(m.group(1))
        if not nm:
            continue
        # 번호에 가장 가까운 상호를 고른다
        dist = abs((max(0, at - 160) + m.start()) - at)
        if best is None or dist < best[0]:
            best = (dist, nm)
    return best[1] if best else None


# ─────────────────────────────────────────────────────────────── 추출 경로
_GOOD = re.compile(r"[가-힣0-9A-Za-z,.\-:()원₩%\s]")


def _garbage(t: str) -> float:
    t = re.sub(r"\s", "", t or "")
    return 1.0 - (sum(1 for c in t if _GOOD.match(c)) / len(t) if t else 0.0)


def _native(path: str):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".pdf":
        import pdfplumber
        words, buf = [], []
        with pdfplumber.open(path) as pdf:
            for pg in pdf.pages[:2]:
                buf.append(pg.extract_text() or "")
                for w in pg.extract_words(x_tolerance=1.5, y_tolerance=2.5) or []:
                    t = (w.get("text") or "").strip()
                    if t:
                        words.append((float(w["x0"]), float(w["x1"]), float(w["top"]),
                                      float(w["bottom"]), 100.0, t))
        return "\n".join(buf), words
    if ext == ".hwp":
        import olefile
        f = olefile.OleFileIO(path)
        comp = True
        try:
            comp = bool(f.openstream("FileHeader").read()[36] & 0x01)
        except Exception:
            pass
        parts = []
        for e in f.listdir():
            if len(e) > 1 and e[0] == "BodyText":
                data = f.openstream(e).read()
                if comp:
                    try:
                        data = zlib.decompress(data, -15)
                    except Exception:
                        continue
                i = 0
                while i + 4 <= len(data):
                    hh = int.from_bytes(data[i:i + 4], "little")
                    tag, size = hh & 0x3FF, (hh >> 20) & 0xFFF
                    i += 4
                    if size == 0xFFF:
                        size = int.from_bytes(data[i:i + 4], "little"); i += 4
                    body, i = data[i:i + size], i + size
                    if tag == 67:
                        s = body.decode("utf-16-le", errors="ignore")
                        parts.append("".join(c for c in s if c == "\n" or ord(c) >= 32))
        f.close()
        return "\n".join(parts), []
    if ext == ".xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        parts = []
        for ws in wb.worksheets[:2]:
            for row in ws.iter_rows(max_row=120, values_only=True):
                parts.extend(str(v) for v in row if v is not None)
        wb.close()
        return "\n".join(parts), []
    return "", []


def _tess(img, args):
    fd, p = tempfile.mkstemp(suffix=".png")
    os.close(fd)
    try:
        ok, buf = cv2.imencode(".png", img)
        buf.tofile(p)
        r = subprocess.run([TESS, p, "stdout"] + args, capture_output=True, timeout=180)
        return r.stdout.decode("utf-8", "ignore")
    finally:
        try:
            os.remove(p)
        except OSError:
            pass


def _words_to_text(words) -> str:
    """TSV 단어를 줄 단위 텍스트로 되돌린다.

    같은 이미지에 `_tess`(텍스트)와 `_tsv`(좌표)를 각각 부르면 tesseract 를 두 번 돌린다.
    TSV 한 번으로 좌표와 텍스트를 다 얻는다 — 호출 수가 절반이 되고, 그만큼 빨라진다.
    """
    return "\n".join(" ".join(w[5] for w in sorted(row, key=lambda w: w[0]))
                     for row in group_rows(words))


def _tsv(img, psm):
    out = _tess(img, ["-l", "kor+eng", "--psm", psm,
                      "-c", "preserve_interword_spaces=1", "tsv"])
    ws = []
    for ln in out.splitlines()[1:]:
        c = ln.split("\t")
        if len(c) < 12 or c[11].strip() in ("", "-1"):
            continue
        try:
            l, t, w, h, cf = int(c[6]), int(c[7]), int(c[8]), int(c[9]), float(c[10])
        except ValueError:
            continue
        ws.append((float(l), float(l + w), float(t), float(t + h), cf, c[11]))
    return ws


def _crop_content(gray):
    """흰 여백을 걷어내고 내용만 남긴다. 작으면 확대한다.

    ⚠ 실제로 걸린 것: 감열지 영수증이 A4 한복판에 **작게** 스캔돼 페이지의 90%가 여백이었다.
       tesseract 가 `7,000` 을 `54.900` 으로 뭉갰다. 잘라내고 확대하면 같은 글자가 제대로 읽힌다.
       (§3 의 '업스케일은 역효과'는 이미 페이지를 꽉 채운 8.7Mpx 스캔 얘기다 — 여기는 반대 상황이다)
    """
    inv = cv2.bitwise_not(gray)
    _, bw = cv2.threshold(inv, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)
    bw = cv2.morphologyEx(bw, cv2.MORPH_OPEN, np.ones((3, 3), np.uint8))
    pts = cv2.findNonZero(bw)
    if pts is None:
        return gray, False
    x, y, w, h = cv2.boundingRect(pts)
    H, W = gray.shape[:2]
    # 내용이 페이지의 60% 미만일 때만 자른다. 이미 꽉 찼으면 건드리지 않는다.
    if w * h >= 0.6 * W * H or w < 80 or h < 80:
        return gray, False
    m = int(0.02 * max(w, h)) + 6
    x0, y0 = max(0, x - m), max(0, y - m)
    x1, y1 = min(W, x + w + m), min(H, y + h + m)
    crop = gray[y0:y1, x0:x1]
    # 잘라낸 뒤 글자가 작으면 확대해서 tesseract 가 볼 수 있게 만든다
    scale = 1.0
    if crop.shape[1] < 1400:
        scale = min(3.0, 1400 / max(1, crop.shape[1]))
        crop = cv2.resize(crop, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
    return crop, True


def _score(text, words) -> tuple:
    """판독 결과의 **확인된 정보량**. 많이 읽은 쪽이 아니라 확인된 게 많은 쪽이 낫다.

    ⚠ 우리 번호는 정보가 아니다 — 어느 증빙에나 찍힌다. **거래처 번호**를 찾았는지를 센다.
    """
    amt = extract_amounts(words) if words else {"산술검증": False, "합계": None}
    brns = find_brns(text)
    theirs = len([b for b in brns if b != OUR_BRN])
    return (bool(amt.get("산술검증")), theirs, len(brns),
            bool(find_date(text)), bool(amt.get("합계")))


def _thin(text, words) -> bool:
    """다시 볼 만큼 부실한가 — 금액이 검산 안 됐거나 **거래처 번호**·일자를 못 찾았을 때.

    실측: 엔캠 세금계산서가 우리 번호만 찾고도 「번호 1개」라 충분으로 판정돼 재시도를
    건너뛰었다. 300dpi 에서 304→904 로 오독한 것이라 400dpi 면 살아난다.
    """
    s = _score(text, words)
    return not (s[0] and s[1] and s[3])


def _scan(path: str, dpi: float = 300.0):
    ext = os.path.splitext(path)[1].lower()
    if ext in (".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp"):
        img = cv2.imdecode(np.fromfile(path, np.uint8), cv2.IMREAD_COLOR)
    else:
        import pypdfium2 as pdfium
        d = pdfium.PdfDocument(path)
        try:
            pil = d[0].render(scale=dpi / 72).to_pil().convert("RGB")
            img = cv2.cvtColor(np.array(pil), cv2.COLOR_RGB2BGR)
        finally:
            d.close()
    g = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    # ① 여백을 걷어낸다(필요하면 확대). 크롭 결과가 멀쩡하면 원본은 아예 안 돌린다 — 속도.
    g_crop, cropped = _crop_content(g)
    if cropped:
        w_crop = _tsv(g_crop, "4")
        t_crop = _words_to_text(w_crop)
        if len(re.sub(r"\s", "", t_crop)) >= 60 and _garbage(t_crop) <= 0.15:
            g, w4, plain = g_crop, w_crop, t_crop
        else:
            w_full = _tsv(g, "4")
            t_full = _words_to_text(w_full)
            if len(re.sub(r"\s", "", t_crop)) > len(re.sub(r"\s", "", t_full)):
                g, w4, plain = g_crop, w_crop, t_crop
            else:
                w4, plain = w_full, t_full
    else:
        w4 = _tsv(g, "4")
        plain = _words_to_text(w4)

    # ② 회전 게이트 — 멀쩡한 문서를 돌리면 오히려 망가진다
    gr0, rot = _garbage(plain), 0
    if gr0 > 0.10:
        m = re.search(r"Rotate:\s*(-?\d+)", _tess(g, ["--psm", "0"]))
        deg = int(m.group(1)) % 360 if m else 0
        code = {90: cv2.ROTATE_90_CLOCKWISE, 180: cv2.ROTATE_180,
                270: cv2.ROTATE_90_COUNTERCLOCKWISE}.get(deg)
        if code is not None:
            g2 = cv2.rotate(g, code)
            w2 = _tsv(g2, "4")
            p2 = _words_to_text(w2)
            if _garbage(p2) < gr0:
                g, w4, plain, rot = g2, w2, p2, deg

    # ③ psm6 을 더 볼 이유가 있으면 한 번 더 — 금액이 검증 안 됐거나 일자·사업자번호가 비었을 때.
    #    실측: 같은 감열지에서 psm4 는 금액줄을, psm6 은 거래일시줄을 살렸다. 서로 보완한다.
    need6 = (not extract_amounts(w4)["산술검증"]
             or find_date(plain) is None
             or not find_brns(plain))
    if not need6:
        return plain, w4, rot
    w6 = _tsv(g, "6")
    t6 = _words_to_text(w6)
    words = w6 if (extract_amounts(w6)["산술검증"]
                   and not extract_amounts(w4)["산술검증"]) else (w4 + w6)
    return plain + "\n" + t6, words, rot


# ─────────────────────────────────────────────────────────────── 진입점
def read(path: str) -> dict | None:
    """extract.read_evidence() 와 같은 스키마. 판독 불가면 None (LLM 이 받는다)."""
    text, words = _native(path)
    rot, 경로 = 0, "native"
    if len(re.sub(r"\s", "", text)) < 20:
        if not (_HAS_TESS and _HAS_CV):
            return None
        text, words, rot = _scan(path)
        경로 = "scan"
        # ★ 300dpi 로 부실하면 400dpi 로 한 번 더 보고 **나은 쪽**을 고른다.
        #   400 을 기본으로 올렸더니 잘 읽히던 90° 회전 스캔이 망가졌다(실측).
        #   그래서 잘 읽힌 문서는 건드리지 않는다. 비용은 실패 건에만 든다.
        if _thin(text, words) and os.environ.get("RND_SCAN_RETRY", "1") != "0":
            try:
                t2, w2, r2 = _scan(path, dpi=400.0)
                if _score(t2, w2) > _score(text, words):
                    text, words, rot = t2, w2, r2
            except Exception as e:
                print(f"[ocr] 400dpi 재시도 실패(무시): {e}", file=sys.stderr)
    if len(re.sub(r"\s", "", text)) < 20:
        return None

    doc, conf, ev = classify_doc(text)
    amt = extract_amounts(words) if words else dict(
        합계=None, 공급가액=None, 세액=None, 방법="좌표없음", 산술검증=False)
    brns = find_brns(text)
    _items_text = extract_items_from_text(text)
    # 카드전표(매출전표)에는 품목이 없다. 그럴싸한 조각을 만들어내느니 비운다.
    _is_card = bool(re.search(r"매\s*출\s*전\s*표|카\s*드\s*전\s*표|신용카드\s*매출|"
                              r"[카|]\s*[드匚]\s*[종총]\s*류", text[:2000], re.I))
    if _items_text:
        _items = [{"품목명": nm, "수량": None, "금액": None, "confidence": 0.7,
                   "note": "품목표에서 추출"} for nm in _items_text]
    elif _is_card:
        _items = []                      # 카드전표 — 문서에 품목이 없다
    else:
        _items = extract_items(words) if words else []

    # 거래 방향 — 추측하지 않는다. 영수증류는 「가맹점=공급자」가 서식으로 보장된다.
    공급자 = {"name": None, "brn": ""}
    공급받는자 = {"name": None, "brn": ""}
    ours = [b for b in brns if OUR_BRN and b == OUR_BRN]
    theirs = [b for b in brns if b != OUR_BRN]
    if doc == "결제영수증":
        # 영수증·카드전표는 가맹점이 곧 공급자다. 서식이 그걸 보장하므로 이름도 붙인다.
        if theirs:
            공급자["brn"] = theirs[0]
        elif brns and not ours:
            공급자["brn"] = brns[0]
        if ours:
            공급받는자["brn"] = OUR_BRN
        # ① 라벨(`상호:`) → ② 사업자번호와 같은 줄 (카드전표는 라벨 없이 같은 줄에 찍는다)
        공급자["name"] = (find_merchant(text)
                        or (find_name_in_brn_row(words, 공급자["brn"])
                            if 공급자["brn"] else None))
    elif ours and theirs:
        # ★ 우리 번호와 **같은 줄**에 찍힌 번호를 먼저 본다. 세금계산서는 두 등록번호를
        #   나란히 찍으므로 이게 가장 확실한 자리 신호다. 하단의 발급대행사 번호와 갈린다.
        공급자["brn"] = brn_paired_with_ours(text) or (theirs[0] if len(theirs) == 1 else "")
        if 공급자["brn"]:
            공급받는자["brn"] = OUR_BRN
            # ★ 상호는 **번호가 찍힌 자리**를 기준으로 집는다. 텍스트 순서로 집으면
            #   좌우 블록이 섞여 우리 회사가 거래처로 들어간다(실측).
            공급자["name"] = find_vendor_name(words, text, 공급자["brn"])
        # ⚠ 이름은 **native(좌표가 살아있는) 경로에서만** 채운다.
        #   세금계산서·계산서는 공급자/공급받는자 블록이 좌우로 나란히 있는데, OCR 로
        #   펼치면 두 블록 글자가 뒤섞인다. 실측: 스캔 세금계산서에서 **우리 회사명이
        #   공급자로** 잡혔다 — 거래처가 뒤집히면 과거 매칭이 통째로 어긋난다(§5.2).
        #   번호는 체크섬으로 확정돼 있으니 방향은 이미 정확하다. 이름만 비운다.
        if 경로 == "native":
            # ★ 자리 기반 결과(find_vendor_name)를 **덮어쓰지 않는다.** 예전에는 여기서
            #   통째로 덮어써서 「성 상호 성」·「김용민 공 ㈜매그나텍」(우리 회사!)이 남았다.
            #   보조로 쓰되 같은 필터를 통과시킨다 — 의심스러우면 비운다.
            if not 공급자.get("name"):
                _k = 공급자.get("brn") or (theirs[0] if theirs else "")
                공급자["name"] = _clean_vendor_name(find_name_near_brn(text, _k) or "") if _k else None
            공급받는자["name"] = find_name_near_brn(text, OUR_BRN)

    return {
        "서류종류": _DOC_MAP.get(doc, "unknown"),
        "공급자": 공급자,
        "공급받는자": 공급받는자,
        "일자": find_date(text),
        "공급가액": amt["공급가액"],
        "세액": amt["세액"],
        "합계": amt["합계"],
        # ① 품목 표(세금계산서·계산서·거래명세서) → ② 좌표 기반(POS 영수증)
        #   품목이 비목을 정한다. 다만 **카드전표에는 품목이 인쇄되지 않는다** —
        #   결제 사실만 적는 서식이라 LLM 으로도 못 읽는다(문서에 없는 정보다).
        #   실측: 그런 전표에서 '신용구매'·'할부' 조각이 품목으로 새어나와
        #   ['구매','할게'] 가 됐고, 그 쓰레기가 비목 모델 입력을 오염시켰다.
        #   **카드전표는 품목을 비우고 가맹점으로 판단하게 둔다** — 그게 유일한 근거다.
        "품목": _items,
        # ── 아래는 로컬 경로 표시용. 기존 소비자는 무시하고, 필요하면 근거로 쓴다.
        "_판독": "로컬OCR",
        "_경로": 경로,
        "_회전": rot,
        "_서식": doc,
        "_서식확신도": conf,
        "_서식근거": ev,
        "_금액방법": amt["방법"],
        "_산술검증": amt["산술검증"],
        "_사업자번호": brns,
        "_OCR품질": round(1 - _garbage(text), 3) if 경로 == "scan" else None,
        # ★ 학습 피처. app.evidence_doc_reads.본문텍스트 로 들어간다.
        #   사람이 비목을 고친 기록(app.decisions)과 이어 붙여 다음 모델을 학습한다.
        "_본문": text[:20000],
    }


def good_enough(r: dict | None) -> bool:
    """LLM 을 안 불러도 되는가.

    ① 산술로 검증된 금액(s+v=t 이고 v≈s/10)이 있으면 인정한다.
    ② 산술이 안 맞아도 **문서가 「합계금액」·「승인금액」이라고 직접 말한 값**(S0)이 있고
       일자·사업자번호까지 읽혔으면 인정한다.

    ②를 넣은 이유: 카드전표는 부가세가 정확히 1/10 이 아닐 때가 있다
    (188,368 → 18,832). 그때 산술검증만 보면 **정확히 읽은 합계를 두고도** LLM 을
    3.5만 토큰씩 부른다. 실측: 147건에서 LLM 호출 25건 → 19건.
    ②로 통과한 건들의 합계는 정답셋에서 정답으로 확인됐다.
    """
    if not (r and r.get("합계")):
        return False
    if r.get("_산술검증"):
        return True
    return bool(str(r.get("_금액방법") or "").startswith("S0")
                and r.get("일자") and (r.get("_사업자번호") or []))
